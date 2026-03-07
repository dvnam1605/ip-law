"""
Merge WIPO Excel downloads → JSON duy nhất, deduplicate theo ST13.
Chuyển 14 cột Excel thành format phù hợp cho Neo4j ingestion.

Usage:
    python -m utils.xlsx_merger
    python -m utils.xlsx_merger --input-dir data/trademarks/vn_excel --output data/trademarks/all_vn_trademarks.json
"""
import json
import logging
import re
from pathlib import Path
from typing import Dict, List

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).parent.parent
DEFAULT_INPUT_DIR = PROJECT_ROOT / "data" / "trademarks" / "vn_excel"
DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "trademarks" / "all_vn_trademarks.json"

# Map Excel headers → field names
COLUMN_MAP = {
    "Brand name": "brand_name",
    "logo": "logo",
    "Status": "status",
    "inids.status_Date": "status_date",
    "IP office": "ip_office",
    "Owner": "owner_name",
    "Registration date": "registration_date",
    "Application number": "application_number",
    "general_words.registrationNumber": "registration_number",
    "Nice classification": "nice_classes",
    "Feature": "feature",
    "Application date": "application_date",
    "Expiry date": "expiry_date",
    "ST13": "st13",
}


def _parse_owner(raw: str) -> tuple:
    """Parse 'Nguyễn Văn A (Việt Nam)' → (name, country)."""
    if not raw:
        return "", ""
    raw = raw.strip()
    if "(" in raw:
        idx = raw.rfind("(")
        name = raw[:idx].strip()
        country = raw[idx:].strip("() ")
        return name, country
    return raw, ""


def _parse_nice_classes(raw) -> List[str]:
    """Parse '9, 11, 35' hoặc '9' → ['9', '11', '35']."""
    if not raw:
        return []
    s = str(raw).strip()
    parts = re.split(r"[,;\s]+", s)
    return [p.strip() for p in parts if p.strip() and p.strip().isdigit()]


def parse_excel_file(xlsx_path: Path) -> List[Dict]:
    """Đọc 1 file Excel WIPO, trả về list of records."""
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)

    if "Export" not in wb.sheetnames:
        logger.warning(f"No 'Export' sheet in {xlsx_path.name}")
        wb.close()
        return []

    ws = wb["Export"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    headers = [str(h).strip() if h else "" for h in rows[0]]

    # Map header index → field name
    col_map = {}
    for i, h in enumerate(headers):
        if h in COLUMN_MAP:
            col_map[i] = COLUMN_MAP[h]

    records = []
    for row in rows[1:]:
        raw = {}
        for i, field in col_map.items():
            val = row[i] if i < len(row) else None
            raw[field] = val

        # Parse owner → name + country
        owner_name, owner_country = _parse_owner(str(raw.get("owner_name", "") or ""))

        # Parse nice classes
        nice_classes = _parse_nice_classes(raw.get("nice_classes"))

        record = {
            "brand_name": str(raw.get("brand_name", "") or "").strip(),
            "owner_name": owner_name,
            "owner_country": owner_country,
            "ip_office": str(raw.get("ip_office", "") or "").strip(),
            "status": str(raw.get("status", "") or "").strip(),
            "status_date": str(raw.get("status_date", "") or "").strip(),
            "registration_number": str(raw.get("registration_number", "") or "").strip(),
            "application_number": str(raw.get("application_number", "") or "").strip(),
            "nice_classes": nice_classes,
            "feature": str(raw.get("feature", "") or "").strip(),
            "registration_date": str(raw.get("registration_date", "") or "").strip(),
            "application_date": str(raw.get("application_date", "") or "").strip(),
            "expiry_date": str(raw.get("expiry_date", "") or "").strip(),
            "st13": str(raw.get("st13", "") or "").strip(),
        }

        # Skip empty records
        if not record["brand_name"] and not record["st13"]:
            continue

        records.append(record)

    return records


def merge_all(input_dir: Path = None, output_path: Path = None) -> List[Dict]:
    """
    Merge tất cả file .xlsx trong input_dir, deduplicate theo st13.
    Trả về list records đã merged.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl required: pip install openpyxl")

    input_dir = input_dir or DEFAULT_INPUT_DIR
    output_path = output_path or DEFAULT_OUTPUT

    xlsx_files = sorted(input_dir.glob("*.xlsx"))
    if not xlsx_files:
        logger.warning(f"No .xlsx files found in {input_dir}")
        return []

    logger.info(f"Merging {len(xlsx_files)} Excel files from {input_dir}...")

    all_records = []
    seen_st13 = set()
    duplicates = 0

    for xlsx_file in xlsx_files:
        records = parse_excel_file(xlsx_file)
        for rec in records:
            st13 = rec.get("st13", "")
            if st13 and st13 in seen_st13:
                duplicates += 1
                continue
            if st13:
                seen_st13.add(st13)
            all_records.append(rec)

        logger.info(f"  {xlsx_file.name}: {len(records)} records")

    # Count VN records
    vn_count = sum(1 for r in all_records if r.get("ip_office") == "VN")

    logger.info(f"\nMerge complete:")
    logger.info(f"  Total records: {len(all_records):,}")
    logger.info(f"  VN records:    {vn_count:,}")
    logger.info(f"  Duplicates:    {duplicates:,}")

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(all_records, f, ensure_ascii=False, indent=2)
    logger.info(f"  Saved to: {output_path}")

    return all_records


# ── CLI ────────────────────────────────────────────────

def main():
    import argparse

    parser = argparse.ArgumentParser(description="Merge WIPO Excel downloads into single JSON")
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_INPUT_DIR)
    parser.add_argument("--output", "-o", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
    merge_all(args.input_dir, args.output)


if __name__ == "__main__":
    main()
